import logging
from datetime import datetime
from django.core.exceptions import PermissionDenied
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Avg
from .models import Predmets, PA, Student, Group, Grade
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, date

from collections import defaultdict

def index(request):
    return render(request, 'index.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if user.role == 'prepod':
                return redirect('prepodovat')
            elif user.role == 'uchebnaya_chast':
                return redirect('uchebnaya')
        else:
            messages.error(request, "Неверное имя пользователя или пароль")

    return render(request, 'index.html')

@login_required
def redirect_view(request):
    print(f"Авторизован пользователь: {request.user.username}")

    return redirect('prepodovat')  # Всех направляем на страницу преподавателя

def logout_view(request):
    logout(request)
    return redirect('index')  # После выхода отправляем на главную


@login_required
def edit_grade(request, pa_id):
    if request.method == "POST":
        # Получаем выбранный предмет, группу и семестр
        selected_predmet = request.POST.get('predmet')
        selected_group = request.POST.get('group')
        selected_semester = request.POST.get('semester')

        pa = get_object_or_404(PA, id=pa_id)
        new_score = request.POST.get('score')
        if new_score:
            pa.score = new_score
            pa.save()
            messages.success(request, "Оценка успешно обновлена.")
        else:
            messages.error(request, "Оценка не может быть пустой.")

        # Перенаправляем на ту же страницу с сохраненными параметрами
        return HttpResponseRedirect(
            f"{reverse('prepodovat')}?group={selected_group}&predmet={selected_predmet}&semester={selected_semester}"
        )

    return redirect('prepodovat')


def get_semester_by_date(input_date, course):
    """
    Определяет семестр на основе даты и курса.
    - С 1 сентября по 31 декабря — 1 семестр.
    - С 1 января по 30 июня — 2 семестр.
    - Курс влияет на нумерацию семестров (1 курс — 1, 2; 2 курс — 3, 4 и т.д.).
    """
    if (input_date.month >= 9 and input_date.month <= 12):
        return (course - 1) * 2 + 1  # Первый семестр
    elif input_date.month >= 1 and input_date.month <= 6:
        return (course - 1) * 2 + 2  # Второй семестр
    else:
        return None  # Летние месяцы (июль, август) не относятся к семестрам

@login_required
def prepodovat_view(request):
    groups = Group.objects.all()
    selected_group = request.GET.get('group')
    selected_predmet = request.GET.get('predmet')
    selected_semester = request.GET.get('semester')  # Получаем выбранный семестр

    students = []
    grades_history = defaultdict(dict)
    dates = set()

    # Получаем привязанного преподавателя (если есть)
    prepod = request.user.prepod if hasattr(request.user, 'prepod') else None

    if selected_group:
        group = Group.objects.get(id=selected_group)
        available_semesters = group.get_available_semesters()  # Доступные семестры
        current_semester = group.get_current_semester_for_group()  # Текущий семестр

        # Ограничиваем выбор семестра только доступными
        if selected_semester:
            selected_semester = int(selected_semester)
            if selected_semester not in available_semesters:
                selected_semester = current_semester  # Если выбран некорректный семестр, используем текущий
        else:
            selected_semester = current_semester  # По умолчанию текущий семестр

        # Фильтруем предметы по преподавателю (если он привязан)
        if prepod:
            predmets = Predmets.objects.filter(group_id=selected_group, name__prepod=prepod).select_related('name')
        else:
            predmets = Predmets.objects.filter(group_id=selected_group).select_related('name')

        if selected_predmet:
            try:
                predmet = Predmets.objects.get(id=selected_predmet, group_id=selected_group)
                students = Student.objects.filter(group_id=selected_group).select_related('group')
                pas = PA.objects.filter(subject=predmet, semester=selected_semester).select_related('student')

                for pa in pas:
                    grades_history[pa.student.id][pa.date] = {'score': pa.score, 'id': pa.id}
                    dates.add(pa.date)

                if request.method == "POST":
                    for student in students:
                        score = request.POST.get(f'score_{student.id}')
                        date_score = request.POST.get(f'date_{student.id}')
                        if score and date_score:
                            try:
                                input_date = date.fromisoformat(date_score)  # Преобразуем строку в дату
                                semester = get_semester_by_date(input_date, group.get_course())  # Определяем семестр
                                if semester is None:
                                    messages.error(request, "Дата не относится к учебному семестру.")
                                    continue

                                PA.objects.create(
                                    student=student,
                                    subject=predmet,
                                    date=date_score,
                                    score=score,
                                    semester=semester  # Используем автоматически определенный семестр
                                )
                            except ValueError:
                                messages.error(request, "Некорректный формат даты.")
                    return HttpResponseRedirect(request.path + f"?group={selected_group}&predmet={selected_predmet}&semester={selected_semester}")

            except ObjectDoesNotExist as e:
                logger.error(f"Ошибка: {e}")
                messages.error(request, "Предмет или группа не найдены")
                return HttpResponseRedirect(request.path)

    # Фильтруем группы по предметам, которые ведет преподаватель (если он привязан)
    if prepod:
        groups = Group.objects.filter(predmets__name__prepod=prepod).distinct()

    dates = sorted(dates)

    return render(request, 'prepodovat.html', {
        'groups': groups,
        'predmets': predmets if selected_group else [],
        'students': students,
        'grades_history': grades_history,
        'dates': dates,
        'selected_group': selected_group,
        'selected_predmet': selected_predmet,
        'selected_semester': selected_semester,
        'current_semester': current_semester if selected_group else None,  # Текущий семестр
        'available_semesters': available_semesters if selected_group else [],  # Доступные семестры
    })


@login_required
def uchebnaya_view(request):
    # Проверяем, что пользователь является учебной частью
    if request.user.role != 'uchebnaya_chast':
        raise PermissionDenied("Доступ запрещен")
    return render(request, 'uchebnaya.html')


@login_required
def dopusk_report_view(request):
    # Проверяем, что пользователь является учебной частью
    if request.user.role != 'uchebnaya_chast':
        raise PermissionDenied("Доступ запрещен")

    groups = Group.objects.all()  # Получаем все группы
    selected_group = request.GET.get('group')
    selected_predmet = request.GET.get('predmet')

    students_data = []  # Список для хранения данных о студентах

    if selected_group:
        # Получаем предметы, которые изучает выбранная группа
        predmets = Predmets.objects.filter(group_id=selected_group).select_related('name')

        if selected_predmet:
            # Получаем предмет и группу
            predmet = Predmets.objects.get(id=selected_predmet, group_id=selected_group)
            students = Student.objects.filter(group_id=selected_group)

            # Собираем количество оценок у каждого студента
            grades_count = {}
            for student in students:
                grades = PA.objects.filter(student=student, subject=predmet).values_list('score', flat=True)
                grades_count[student.id] = len(grades)

            # Определяем максимальное количество оценок в группе
            max_grades_count = max(grades_count.values(), default=0)

            # Анализируем оценки каждого студента
            for student in students:
                # Получаем все оценки студента по выбранному предмету
                grades = PA.objects.filter(student=student, subject=predmet).values_list('score', flat=True)

                # Проверяем, есть ли среди оценок "2"
                has_failing_grade = 2 in grades

                # Проверяем, есть ли пропущенные оценки (количество оценок меньше максимального в группе)
                has_missing_grades = len(grades) < max_grades_count

                # Формируем данные для отчета
                students_data.append({
                    'name': student.name,
                    'grades': list(grades),  # Все оценки студента
                    'is_allowed': not has_failing_grade and not has_missing_grades,  # Допущен, если нет "2" и количество оценок совпадает
                })

    return render(request, 'dopusk_report.html', {
        'groups': groups,
        'predmets': predmets if selected_group else [],
        'students_data': students_data,
        'selected_group': selected_group,
        'selected_predmet': selected_predmet,
    })

@login_required
def uspevaemost_report_view(request):
    # Логика для анализа успеваемости
    return render(request, 'uspevaemost_report.html')


@login_required
def generate_dopusk_report(request):
    # Получаем выбранный предмет и группу из GET-параметров
    selected_predmet = request.GET.get('predmet')
    selected_group = request.GET.get('group')

    if not selected_predmet or not selected_group:
        return HttpResponse("Предмет или группа не выбраны", status=400)

    # Получаем данные о предмете и группе
    try:
        predmet = Predmets.objects.get(id=selected_predmet, group_id=selected_group)
        students = Student.objects.filter(group_id=selected_group)
    except Predmets.DoesNotExist:
        return HttpResponse("Предмет или группа не найдены", status=404)

    # Собираем количество оценок у каждого студента
    grades_count = {}
    for student in students:
        grades = PA.objects.filter(student=student, subject=predmet).values_list('score', flat=True)
        grades_count[student.id] = len(grades)

    # Определяем максимальное количество оценок в группе
    max_grades_count = max(grades_count.values(), default=0)

    # Создаем новый Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по допуску"

    # Заголовок отчета
    ws['B2'] = "Отчет по допуску к экзамену"
    ws['B2'].font = Font(bold=True, size=14)
    ws.merge_cells('B2:D2')  # Объединяем ячейки для заголовка

    # Информация о предмете и группе
    ws['A3'] = "Предмет:"
    ws['B3'] = predmet.name.name  # Извлекаем название предмета
    ws['A4'] = "Группа:"
    ws['B4'] = predmet.group.name  # Извлекаем название группы

    # Заголовки таблицы
    ws['A6'] = "ФИО студента"
    ws['B6'] = "Оценки"
    ws['C6'] = "Допуск к экзамену"
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Заполняем таблицу данными
    row = 7
    for student in students:
        grades = PA.objects.filter(student=student, subject=predmet).values_list('score', flat=True)
        has_failing_grade = 2 in grades
        has_missing_grades = len(grades) < max_grades_count

        ws[f'A{row}'] = student.name
        ws[f'B{row}'] = ", ".join(map(str, grades))
        ws[f'C{row}'] = "Допущен" if not has_failing_grade and not has_missing_grades else "Не допущен"
        row += 1

    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Сохраняем файл в HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="dopusk_report.xlsx"'
    wb.save(response)

    return response


@login_required
def zadolzhennosti_report_view(request):
    # Проверяем, что пользователь является учебной частью
    if request.user.role != 'uchebnaya_chast':
        raise PermissionDenied("Доступ запрещен")

    # Получаем выбранный предмет и группу из GET-параметров
    selected_predmet = request.GET.get('predmet')
    selected_group = request.GET.get('group')

    students_data = []
    predmets = Predmets.objects.select_related('group', 'name').all()
    groups = Group.objects.all()

    if selected_group:
        # Получаем студентов выбранной группы
        students = Student.objects.filter(group_id=selected_group)

        if selected_predmet:
            # Отчет по задолженностям для конкретного предмета и группы
            predmet = Predmets.objects.get(id=selected_predmet, group_id=selected_group)

            # Получаем все даты, за которые есть оценки у студентов по этому предмету
            all_dates = PA.objects.filter(subject=predmet).values_list('date', flat=True).distinct()
            all_dates = sorted(all_dates)  # Сортируем даты

            for student in students:
                # Получаем все оценки "2" для студента по выбранному предмету
                failing_grades = PA.objects.filter(student=student, subject=predmet, score=2)
                failing_dates = [grade.date.strftime("%Y-%m-%d") for grade in failing_grades]

                # Получаем все даты, за которые у студента есть оценки
                student_dates = PA.objects.filter(student=student, subject=predmet).values_list('date', flat=True)
                student_dates = set(student_dates)

                # Находим даты, за которые у студента нет оценок
                missing_dates = [date.strftime("%Y-%m-%d") for date in all_dates if date not in student_dates]

                # Формируем данные для отчета
                students_data.append({
                    'name': student.name,
                    'failing_dates': failing_dates,  # Даты с оценкой "2"
                    'missing_dates': missing_dates,  # Даты с отсутствующими оценками
                })
        else:
            # Отчет по задолженностям для конкретной группы (по всем предметам)
            for student in students:
                # Получаем все оценки "2" для студента
                failing_grades = PA.objects.filter(student=student, score=2).select_related('subject')
                failing_dates = [f"{grade.subject.name.name} ({grade.date.strftime('%Y-%m-%d')})" for grade in failing_grades]

                # Получаем все предметы, которые изучает группа
                group_predmets = Predmets.objects.filter(group_id=selected_group)

                # Находим даты, за которые у студента нет оценок по каждому предмету
                missing_dates = []
                for predmet in group_predmets:
                    all_dates = PA.objects.filter(subject=predmet).values_list('date', flat=True).distinct()
                    student_dates = PA.objects.filter(student=student, subject=predmet).values_list('date', flat=True)
                    missing_dates.extend([
                        f"{predmet.name.name} ({date.strftime('%Y-%m-%d')})"
                        for date in all_dates if date not in student_dates
                    ])

                # Формируем данные для отчета
                students_data.append({
                    'name': student.name,
                    'failing_dates': failing_dates,  # Даты с оценкой "2"
                    'missing_dates': missing_dates,  # Даты с отсутствующими оценками
                })

    return render(request, 'zadolzhennosti_report.html', {
        'predmets': predmets,
        'groups': groups,
        'students_data': students_data,
        'selected_predmet': selected_predmet,
        'selected_group': selected_group,
    })


@login_required
def generate_zadolzhennosti_excel(request):
    # Получаем выбранный предмет и группу из GET-параметров
    selected_predmet = request.GET.get('predmet')
    selected_group = request.GET.get('group')

    if not selected_group:
        return HttpResponse("Группа не выбрана", status=400)

    # Получаем студентов выбранной группы
    students = Student.objects.filter(group_id=selected_group)

    # Создаем новый Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Задолженности"

    # Заголовок отчета
    ws['B2'] = "Отчет по задолженностям студентов"
    ws['B2'].font = Font(bold=True, size=14)
    ws.merge_cells('B2:D2')  # Объединяем ячейки для заголовка

    # Информация о группе
    group = Group.objects.get(id=selected_group)
    ws['A3'] = "Группа:"
    ws['B3'] = group.name

    # Информация о предмете (если выбран)
    if selected_predmet:
        predmet = Predmets.objects.get(id=selected_predmet, group_id=selected_group)
        ws['A4'] = "Предмет:"
        ws['B4'] = predmet.name.name

    # Заголовки таблицы
    ws['A6'] = "ФИО студента"
    ws['B6'] = "Задолженности"
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Заполняем таблицу данными
    row = 7
    for student in students:
        if selected_predmet:
            # Отчет по конкретному предмету
            failing_grades = PA.objects.filter(student=student, subject_id=selected_predmet, score=2)
            missing_dates = []
            all_dates = PA.objects.filter(subject_id=selected_predmet).values_list('date', flat=True).distinct()
            student_dates = PA.objects.filter(student=student, subject_id=selected_predmet).values_list('date', flat=True)
            missing_dates = [date.strftime("%Y-%m-%d") for date in all_dates if date not in student_dates]

            debts = []
            if failing_grades.exists():
                debts.extend([f"2 ({grade.date.strftime('%Y-%m-%d')})" for grade in failing_grades])
            if missing_dates:
                debts.extend([f"Отсутствует ({date})" for date in missing_dates])
        else:
            # Отчет по всем предметам
            failing_grades = PA.objects.filter(student=student, score=2).select_related('subject')
            missing_dates = []
            group_predmets = Predmets.objects.filter(group_id=selected_group)
            for predmet in group_predmets:
                all_dates = PA.objects.filter(subject=predmet).values_list('date', flat=True).distinct()
                student_dates = PA.objects.filter(student=student, subject=predmet).values_list('date', flat=True)
                missing_dates.extend([
                    f"Отсутствует {predmet.name.name} ({date.strftime('%Y-%m-%d')})"
                    for date in all_dates if date not in student_dates
                ])

            debts = []
            if failing_grades.exists():
                debts.extend([f"2 {grade.subject.name.name} ({grade.date.strftime('%Y-%m-%d')})" for grade in failing_grades])
            if missing_dates:
                debts.extend(missing_dates)

        ws[f'A{row}'] = student.name
        ws[f'B{row}'] = "; ".join(debts) if debts else "Долгов нет"
        row += 1

    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60

    # Сохраняем файл в HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="zadolzhennosti_report.xlsx"'
    wb.save(response)

    return response

@login_required
def uspevaemost_report_view(request):
    # Проверяем, что пользователь является учебной частью
    if request.user.role != 'uchebnaya_chast':
        raise PermissionDenied("Доступ запрещен")

    # Получаем выбранный предмет и группу из GET-параметров
    selected_predmet = request.GET.get('predmet')
    selected_group = request.GET.get('group')

    students_data = []
    predmets = Predmets.objects.select_related('group', 'name').all()
    groups = Group.objects.all()

    if selected_group:
        # Получаем студентов выбранной группы
        students = Student.objects.filter(group_id=selected_group)

        if selected_predmet:
            # Отчет по конкретному предмету
            predmet = Predmets.objects.get(id=selected_predmet, group_id=selected_group)
            for student in students:
                # Получаем все оценки студента по выбранному предмету
                grades = PA.objects.filter(student=student, subject=predmet).values_list('score', flat=True)
                if grades:
                    average_score = round(sum(grades) / len(grades), 2)
                    failing_grades = PA.objects.filter(student=student, subject=predmet, score=2).count()
                else:
                    average_score = None
                    failing_grades = 0
                students_data.append({
                    'name': student.name,
                    'average_score': average_score,
                    'failing_grades': failing_grades,
                })
        else:
            # Отчет по всем предметам
            for student in students:
                # Получаем все оценки студента
                grades = PA.objects.filter(student=student).values_list('score', flat=True)
                if grades:
                    average_score = round(sum(grades) / len(grades), 2)
                    failing_grades = PA.objects.filter(student=student, score=2).count()
                else:
                    average_score = None
                    failing_grades = 0
                students_data.append({
                    'name': student.name,
                    'average_score': average_score,
                    'failing_grades': failing_grades,
                })

    return render(request, 'uspevaemost_report.html', {
        'predmets': predmets,
        'groups': groups,
        'students_data': students_data,
        'selected_predmet': selected_predmet,
        'selected_group': selected_group,
    })


@login_required
def generate_uspevaemost_excel(request):
    # Получаем выбранный предмет и группу из GET-параметров
    selected_predmet = request.GET.get('predmet')
    selected_group = request.GET.get('group')

    if not selected_group:
        return HttpResponse("Группа не выбрана", status=400)

    # Получаем студентов выбранной группы
    students = Student.objects.filter(group_id=selected_group)

    # Создаем новый Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Успеваемость"

    # Заголовок отчета
    ws['B2'] = "Отчет по успеваемости студентов"
    ws['B2'].font = Font(bold=True, size=14)
    ws.merge_cells('B2:D2')  # Объединяем ячейки для заголовка

    # Информация о группе
    group = Group.objects.get(id=selected_group)
    ws['A3'] = "Группа:"
    ws['B3'] = group.name

    # Информация о предмете (если выбран)
    if selected_predmet:
        predmet = Predmets.objects.get(id=selected_predmet, group_id=selected_group)
        ws['A4'] = "Предмет:"
        ws['B4'] = predmet.name.name

    # Заголовки таблицы
    ws['A6'] = "ФИО студента"
    ws['B6'] = "Средний балл"
    ws['C6'] = "Количество оценок '2'"
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Заполняем таблицу данными
    row = 7
    for student in students:
        if selected_predmet:
            # Отчет по конкретному предмету
            grades = PA.objects.filter(student=student, subject_id=selected_predmet).values_list('score', flat=True)
            if grades:
                average_score = round(sum(grades) / len(grades), 2)
                failing_grades = PA.objects.filter(student=student, subject_id=selected_predmet, score=2).count()
            else:
                average_score = None
                failing_grades = 0
        else:
            # Отчет по всем предметам
            grades = PA.objects.filter(student=student).values_list('score', flat=True)
            if grades:
                average_score = round(sum(grades) / len(grades), 2)
                failing_grades = PA.objects.filter(student=student, score=2).count()
            else:
                average_score = None
                failing_grades = 0

        ws[f'A{row}'] = student.name
        ws[f'B{row}'] = average_score if average_score is not None else "Нет данных"
        ws[f'C{row}'] = failing_grades
        row += 1

    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Сохраняем файл в HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="uspevaemost_report.xlsx"'
    wb.save(response)

    return response


@login_required
def administ_view(request):
    # Проверяем, что пользователь является учебной частью
    if request.user.role != 'uchebnaya_chast':
        raise PermissionDenied("Доступ запрещен")

    groups = Group.objects.all()
    students = Student.objects.select_related('group').all()
    return render(request, 'administ.html', {
        'groups': groups,
        'students': students,
    })

@login_required
def add_student(request):
    if request.method == 'POST':
        name = request.POST.get('student_name')
        group_id = request.POST.get('student_group')
        if name and group_id:
            group = Group.objects.get(id=group_id)
            Student.objects.create(name=name, group=group)
            messages.success(request, 'Студент успешно добавлен.')
        else:
            messages.error(request, 'Ошибка при добавлении студента.')
    return redirect('administ')


@login_required
def delete_student(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_to_delete')
        if student_id:
            student = Student.objects.get(id=student_id)
            student.delete()
            messages.success(request, 'Студент успешно удален.')
        else:
            messages.error(request, 'Ошибка при удалении студента.')
    return redirect('administ')


# Настройка логгера
logger = logging.getLogger(__name__)

@login_required
def import_grades(request):
    if request.method == 'POST':
        # Получаем выбранный предмет, группу и семестр
        selected_predmet = request.POST.get('predmet')
        selected_group = request.POST.get('group')
        selected_semester = request.POST.get('semester')

        if not selected_predmet or not selected_group:
            messages.error(request, "Предмет или группа не выбраны.")
            return redirect('prepodovat')

        # Получаем загруженный файл
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Файл не загружен.")
            return redirect('prepodovat')

        try:
            # Загружаем Excel-файл
            wb = load_workbook(excel_file)
            ws = wb.active

            # Получаем группу и предмет
            group = Group.objects.get(id=selected_group)
            predmet = Predmets.objects.get(id=selected_predmet, group_id=selected_group)
            course = group.get_course()  # Получаем курс группы

            # Получаем даты из заголовков столбцов (начиная со второго столбца)
            dates = [cell.value for cell in ws[1][1:]]  # Первая строка — заголовки

            # Обрабатываем строки файла (начиная со второй строки)
            for row in ws.iter_rows(min_row=2, values_only=True):
                student_name = row[0]  # Первый столбец — ФИО студента
                grades = row[1:]       # Остальные столбцы — оценки по датам

                # Находим студента по ФИО
                student = Student.objects.filter(name=student_name, group_id=selected_group).first()
                if not student:
                    messages.warning(request, f"Студент {student_name} не найден.")
                    continue

                # Сохраняем оценки для каждой даты
                for date_value, score in zip(dates, grades):
                    if not date_value or not score:  # Пропускаем пустые ячейки
                        continue

                    # Проверяем, что оценка в допустимом диапазоне
                    if not (2 <= score <= 5):
                        messages.warning(request, f"Некорректная оценка: {score} для студента {student_name}")
                        continue

                    # Преобразуем дату в строку, если это объект datetime
                    if isinstance(date_value, datetime):
                        date_str = date_value.strftime("%Y-%m-%d")
                    else:
                        try:
                            # Если дата передана как строка, проверяем её формат
                            datetime.strptime(date_value, "%Y-%m-%d")
                            date_str = date_value
                        except ValueError:
                            messages.warning(request, f"Некорректный формат даты: {date_value}")
                            continue

                    # Определяем семестр на основе даты
                    input_date = date.fromisoformat(date_str)
                    semester = get_semester_by_date(input_date, course)  # Используем функцию для определения семестра
                    if semester is None:
                        messages.warning(request, f"Дата {date_str} не относится к учебному семестру.")
                        continue

                    # Сохраняем оценку
                    PA.objects.create(
                        student=student,
                        subject=predmet,
                        date=date_str,
                        score=score,
                        semester=semester  # Используем автоматически определенный семестр
                    )

            messages.success(request, "Оценки успешно импортированы.")
        except Exception as e:
            logger.error(f"Ошибка при импорте оценок: {e}")
            messages.error(request, f"Ошибка при импорте оценок: {e}")

        # Перенаправляем на ту же страницу с сохраненными параметрами
        return HttpResponseRedirect(
            f"{reverse('prepodovat')}?group={selected_group}&predmet={selected_predmet}&semester={selected_semester}"
        )

    return redirect('prepodovat')


def itogi_view(request):
    groups = Group.objects.all()
    selected_group = request.GET.get('group')
    selected_semester = request.GET.get('semester')

    students_data = {}
    subjects = set()
    period_label = ""

    if selected_group:
        group = Group.objects.get(id=selected_group)
        available_semesters = group.get_available_semesters()  # Получаем доступные семестры для группы
        current_semester = group.get_current_semester_for_group()  # Текущий семестр

        if selected_semester:
            students = Student.objects.filter(group_id=selected_group)
            for student in students:
                student_grades = {}
                if selected_semester == "year":
                    # Итоговые оценки за весь год (все семестры)
                    grades = PA.objects.filter(student=student)
                    period_label = "Итоги за год"
                else:
                    # Итоговые оценки за выбранный семестр
                    grades = PA.objects.filter(student=student, semester=selected_semester)
                    if selected_semester == str(current_semester):
                        period_label = f"Итоги за {selected_semester} семестр (Текущий)"
                    elif int(selected_semester) > current_semester:
                        period_label = f"Итоги за {selected_semester} семестр (Следующий)"
                    else:
                        period_label = f"Итоги за {selected_semester} семестр (Предыдущий)"

                # Группируем оценки по предметам и вычисляем средний балл
                for grade in grades:
                    subject_name = grade.subject.name.name
                    if subject_name not in student_grades:
                        student_grades[subject_name] = []
                    student_grades[subject_name].append(grade.score)

                # Вычисляем средний балл для каждого предмета
                for subject, scores in student_grades.items():
                    student_grades[subject] = round(sum(scores) / len(scores), 2)
                    subjects.add(subject)  # Добавляем предмет в общий список

                students_data[student.name] = student_grades

    # Преобразуем множество предметов в список для удобства
    subjects = sorted(subjects)

    return render(request, 'itogi.html', {
        'groups': groups,
        'selected_group': selected_group,
        'selected_semester': selected_semester,
        'available_semesters': available_semesters if selected_group else [],
        'students_data': students_data,
        'subjects': subjects,
        'period_label': period_label,
        'current_semester': current_semester if selected_group else None,
    })

